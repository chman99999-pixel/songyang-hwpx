#!/usr/bin/env python3
"""HWPX 원본 파일의 텍스트만 교체하는 송영서비스 문서 생성기

원본 HWPX를 unpack → section0.xml의 <hp:t> 텍스트만 교체 → pack
레이아웃이 100% 동일하게 유지됩니다.

송영정보 엑셀: 이용자별 오전/오후 송영 시간 (방향별 per-trip 분)
3월 송영서비스 엑셀: 이용자별 서비스 일수, 날짜, 그룹, 단가
"""
import os
import re
import tempfile
from datetime import datetime
from zipfile import ZipFile, ZIP_DEFLATED, ZIP_STORED

from lxml import etree
import openpyxl


def parse_trip_info(filepath):
    """송영정보 엑셀 파싱 - 이용자별 오전/오후 송영 시간(분)"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    # 헤더 찾기
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=10):
        for cell in row:
            if cell.value and '이용자' in str(cell.value):
                header_row = cell.row
                break
        if header_row:
            break

    if not header_row:
        print("  경고: 송영정보 헤더를 찾을 수 없습니다.")
        return {}

    trip_info = {}
    for row in ws.iter_rows(min_row=header_row + 1):
        vals = [c.value for c in row]
        if len(vals) < 4 or not vals[1]:
            continue
        name = str(vals[1]).strip()
        if not name or '총' in name:
            continue

        am_str = str(vals[2]).strip() if vals[2] else '-'
        pm_str = str(vals[3]).strip() if vals[3] else '-'

        # "30분" → 30, "-" → None
        am_min = None
        pm_min = None
        m = re.match(r'(\d+)분?', am_str)
        if m:
            am_min = int(m.group(1))
        m = re.match(r'(\d+)분?', pm_str)
        if m:
            pm_min = int(m.group(1))

        trip_info[name] = {'am_min': am_min, 'pm_min': pm_min}

    return trip_info


def parse_excel(filepath, trip_info=None):
    """엑셀 파일에서 송영서비스 데이터 파싱 (송영정보 반영)"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    header_row = None
    headers = {}
    for row in ws.iter_rows(min_row=1, max_row=10):
        for cell in row:
            if cell.value and '서비스유형' in str(cell.value):
                header_row = cell.row
                break
        if header_row:
            break

    for cell in ws[header_row]:
        if cell.value:
            headers[str(cell.value).strip()] = cell.column - 1

    users = {}
    month = None

    settlement_col = headers.get('결제구분', -1)
    time_col = headers.get('총시간', -1)

    def parse_time_str(s):
        """총시간 파싱: '00100'→60분, '00030'→30분"""
        s = str(s or '').strip()
        if not s or len(s) < 5:
            return 0
        h = int(s[:3]) if s[:3].isdigit() else 0
        m = int(s[3:]) if s[3:].isdigit() else 0
        return h * 60 + m

    for row in ws.iter_rows(min_row=header_row + 1):
        vals = [cell.value for cell in row]
        svc_col = headers.get('서비스유형', -1)
        if svc_col < 0 or svc_col >= len(vals) or str(vals[svc_col]).strip() != '송영서비스':
            continue
        if settlement_col >= 0 and settlement_col < len(vals) and str(vals[settlement_col] or '').strip() == '승인취소':
            continue

        name_col = headers.get('대상자명', -1)
        name = str(vals[name_col]).strip() if name_col >= 0 and vals[name_col] else ''
        if not name:
            continue

        if name not in users:
            group_col = headers.get('대상자 인원', -1)
            group_str = str(vals[group_col]).strip() if group_col >= 0 and vals[group_col] else ''
            users[name] = {'name': name, 'group_str': group_str, 'dates': [], 'row_times': []}

        u = users[name]

        # 총시간 파싱
        row_min = parse_time_str(vals[time_col]) if time_col >= 0 and time_col < len(vals) else 0

        date_col = headers.get('승인일시', -1)
        if date_col >= 0 and vals[date_col]:
            d = vals[date_col]
            if isinstance(d, datetime):
                u['dates'].append(d)
                u['row_times'].append({'date': d, 'min': row_min})
                if not month:
                    month = d.month
            elif isinstance(d, str):
                try:
                    d = datetime.strptime(d.strip()[:10], '%Y-%m-%d')
                    u['dates'].append(d)
                    u['row_times'].append({'date': d, 'min': row_min})
                    if not month:
                        month = d.month
                except:
                    pass

    def fmt_hours(minutes):
        h = minutes // 60
        m = minutes % 60
        if h > 0 and m > 0:
            return f"{h}시간{m}분"
        elif h > 0:
            return f"{h}시간"
        elif m > 0:
            return f"{m}분"
        return "0시간"

    def make_date_range(dates):
        if not dates:
            return ''
        dates = sorted(dates)
        if len(dates) == 1:
            return dates[0].strftime('%m.%d')
        return f"{dates[0].strftime('%m.%d')}~{dates[-1].strftime('%m.%d')}"

    # 계산
    for name, u in users.items():
        gs = u['group_str']
        if '집중' in gs or '1인' in gs:
            u['unit_price'] = 25910
        elif '2인' in gs:
            u['unit_price'] = 17270
        else:
            u['unit_price'] = 13820

        u['dates'].sort()
        u['day_count'] = len(u['dates'])
        u['date_range'] = make_date_range(u['dates'])

        # 송영정보에서 방향별 시간 가져오기
        ti = (trip_info or {}).get(name, {})
        am_min = ti.get('am_min')
        pm_min = ti.get('pm_min')

        # 송영정보에 없는 이용자: 바우처 데이터에서 추론
        if not am_min and not pm_min and u['row_times']:
            times = sorted([rt['min'] for rt in u['row_times']])
            typical = times[len(times) // 2]
            if typical >= 60:
                am_min = 30
                pm_min = 30
            else:
                am_min = typical or 30

        u['am_min'] = am_min
        u['pm_min'] = pm_min

        # 방향별 일수 계산: 총시간으로 추가 방향/부분일 감지
        am_dates = []
        pm_dates = []
        exc_dates = []  # 3번줄용 (오전≠오후인 이용자의 별도일)
        exc_per_min = 30  # 3번줄 per-trip 분 기본값
        detected_extra_min = None

        for rt in u['row_times']:
            actual = rt['min']
            if am_min and pm_min:
                # 양방향 이용자
                expected = am_min + pm_min
                if actual >= expected:
                    # 정상: 양쪽 모두
                    am_dates.append(rt['date'])
                    pm_dates.append(rt['date'])
                elif am_min != pm_min and actual != am_min and actual != pm_min:
                    # 오전≠오후이고, 어느 쪽에도 안 맞는 경우 → 3번줄
                    exc_dates.append(rt['date'])
                    exc_per_min = actual
                elif am_min != pm_min:
                    # 오전 또는 오후 한쪽만 해당
                    if abs(actual - am_min) <= abs(actual - pm_min):
                        am_dates.append(rt['date'])
                    else:
                        pm_dates.append(rt['date'])
                else:
                    # 오전=오후 시간 동일, 구분 불가 → 오전으로 할당
                    am_dates.append(rt['date'])
            elif am_min and not pm_min:
                am_dates.append(rt['date'])
                if actual > am_min:
                    pm_dates.append(rt['date'])
                    if detected_extra_min is None:
                        detected_extra_min = actual - am_min
            elif pm_min and not am_min:
                pm_dates.append(rt['date'])
                if actual > pm_min:
                    am_dates.append(rt['date'])
                    if detected_extra_min is None:
                        detected_extra_min = actual - pm_min

        # 추가 방향 per-trip 시간 설정
        if not pm_min and pm_dates:
            u['pm_min'] = detected_extra_min or am_min
            pm_min = u['pm_min']
        if not am_min and am_dates:
            u['am_min'] = detected_extra_min or pm_min
            am_min = u['am_min']

        u['am_day_count'] = len(am_dates)
        u['pm_day_count'] = len(pm_dates)
        u['am_date_range'] = make_date_range(am_dates)
        u['pm_date_range'] = make_date_range(pm_dates)

        # 3번줄 데이터 (오전≠오후인 이용자의 별도일)
        u['exc_day_count'] = len(exc_dates)
        u['exc_per_min'] = exc_per_min
        u['exc_date_range'] = make_date_range(exc_dates)
        exc_total = exc_per_min * len(exc_dates)
        u['exc_hours_str'] = fmt_hours(exc_total) if exc_dates else ''

        # 방향별 총 분 계산 (표시용)
        am_total_min = (am_min or 0) * u['am_day_count']
        pm_total_min = (pm_min or 0) * u['pm_day_count']

        # 실제 총시간 합계로 비용 계산 (가장 정확)
        actual_total_min = sum(rt['min'] for rt in u['row_times'])

        u['am_hours_raw'] = am_total_min
        u['pm_hours_raw'] = pm_total_min
        u['am_hours_str'] = fmt_hours(am_total_min) if u['am_day_count'] > 0 and am_min else None
        u['pm_hours_str'] = fmt_hours(pm_total_min) if u['pm_day_count'] > 0 and pm_min else None

        # 총 이용시간 (실제 합계 기반)
        u['total_hours_int'] = actual_total_min // 60
        u['total_remain_min'] = actual_total_min % 60
        u['total_time_str'] = fmt_hours(actual_total_min)
        # 실제 분 기준 비용 계산, 10원 단위 절사
        u['final_cost'] = (actual_total_min * u['unit_price'] // 60 // 10) * 10

    return users, month or datetime.now().month


def replace_texts_in_section(section_path, users_data, new_month):
    """section0.xml에서 이용자별 텍스트 교체 (방향별 처리)"""
    tree = etree.parse(section_path)
    root = tree.getroot()
    ns = {'hp': 'http://www.hancom.co.kr/hwpml/2011/paragraph',
          'hs': 'http://www.hancom.co.kr/hwpml/2011/section'}

    mm = f"{new_month:02d}"
    all_t = root.findall('.//hp:t', ns)

    # 이전달 월 파악
    old_month_str = None
    for t_el in all_t:
        if t_el.text and '년' in t_el.text and '월' in t_el.text:
            m = re.search(r'(\d{2})월', t_el.text)
            if m:
                old_month_str = m.group(1)
                break

    if not old_month_str:
        print("  경고: 이전 달 월 정보를 찾을 수 없습니다.")
        return

    old_date_pattern = re.compile(rf'{old_month_str}\.\d{{2}}~{old_month_str}\.\d{{2}}')
    print(f"  이전 달: {old_month_str}월 → 새 달: {mm}월")

    # 방향 추적을 위한 상태
    current_user = None
    current_direction = None  # 'am' or 'pm'
    clearing_row = False  # 해당 방향 없을 때 행 전체 비우기
    direction_row_count = 0  # 현재 이용자의 방향 행 카운트 (1=첫째행, 2=둘째행)

    i = 0
    while i < len(all_t):
        t_el = all_t[i]
        txt = t_el.text or ''
        stripped = txt.strip()

        # 1. 월 교체 (모든 페이지에 적용 - 이용자 유무 무관)
        if '년' in txt and '월' in txt and old_month_str in txt:
            t_el.text = txt.replace(f'{old_month_str}월', f'{mm}월')
            i += 1
            continue

        # 이용자명 발견
        if stripped == '이용자명':
            current_direction = None
            clearing_row = False
            direction_row_count = 0
            for j in range(i + 1, min(i + 10, len(all_t))):
                nt = (all_t[j].text or '').strip()
                if nt and nt != '그룹유형/급여유형' and '/' not in nt and 2 <= len(nt) <= 5:
                    current_user = nt
                    break
            i += 1
            continue

        if not current_user or current_user not in users_data:
            i += 1
            continue

        u = users_data[current_user]

        # am_day_count/pm_day_count 기반 판정 (추가 방향 감지 반영)
        is_pm_only = (u.get('am_day_count', 0) == 0 and u.get('pm_day_count', 0) > 0)
        is_am_only = (u.get('am_day_count', 0) > 0 and u.get('pm_day_count', 0) == 0)

        if stripped in ['주거지→', '거주지→', '제공기관→']:
            direction_row_count += 1
            detected_dir = 'am' if stripped in ['주거지→', '거주지→'] else 'pm'

            if is_pm_only:
                if direction_row_count == 1:
                    # pm-only 첫째행: pm 데이터 채움, 구분을 "제공기관→주거지"로
                    current_direction = 'pm'
                    clearing_row = False
                    if detected_dir == 'am':
                        t_el.text = '제공기관→'
                        for j in range(i + 1, min(i + 3, len(all_t))):
                            nt = (all_t[j].text or '').strip()
                            if nt == '제공기관':
                                all_t[j].text = '주거지'
                                break
                else:
                    # pm-only 둘째행: 비움
                    current_direction = detected_dir
                    clearing_row = True
                    t_el.text = ''
            elif is_am_only:
                if direction_row_count == 1:
                    # am-only 첫째행: am 데이터 채움, 구분을 "주거지→제공기관"으로
                    current_direction = 'am'
                    clearing_row = False
                    if detected_dir == 'pm':
                        t_el.text = '주거지→'
                        for j in range(i + 1, min(i + 3, len(all_t))):
                            nt = (all_t[j].text or '').strip()
                            if nt == '주거지':
                                all_t[j].text = '제공기관'
                                break
                else:
                    # am-only 둘째행: 비움
                    current_direction = detected_dir
                    clearing_row = True
                    t_el.text = ''
            else:
                # 양방향 이용자
                if direction_row_count <= 2:
                    current_direction = detected_dir
                    clearing_row = False
                elif direction_row_count == 3:
                    if u.get('exc_day_count', 0) > 0:
                        current_direction = 'exc'
                        clearing_row = False
                    else:
                        current_direction = detected_dir
                        clearing_row = True
                        t_el.text = ''
                else:
                    current_direction = detected_dir
                    clearing_row = True
                    t_el.text = ''
        elif '총 송영서비스 이용시간' in stripped:
            current_direction = 'total'
            clearing_row = False
        # 행 번호(1,2,3,4)나 섹션 헤더 만나면 clearing 중지
        elif re.match(r'^[1-4]$', stripped):
            clearing_row = False
        elif stripped in ['송영서비스 제공 내역', '송영서비스 시간 및 이용',
                          '연번', '보호자']:
            clearing_row = False

        # clearing_row 모드: 해당 방향이 없으면 장소 등 모든 데이터 비움
        if clearing_row and current_direction in ('am', 'pm', 'exc'):
            structural = ['1', '2', '3', '4', '연번', '구분', '일시 및 시간',
                          '장소', '산출내역', '송영시간', '총 송영서비스 이용시간',
                          '송영서비스 시간 및 이용', '송영서비스 제공 내역']
            if stripped and stripped not in structural and not stripped.startswith('※'):
                t_el.text = ''
                i += 1
                continue

        # 현재 방향에 맞는 per-trip 분, 일수, 날짜범위, 시간문자열
        if current_direction == 'am':
            dir_min = u.get('am_min')
            dir_hours_str = u.get('am_hours_str', '')
            dir_days = u.get('am_day_count', 0)
            dir_date_range = u.get('am_date_range', '')
        elif current_direction == 'pm':
            dir_min = u.get('pm_min')
            dir_hours_str = u.get('pm_hours_str', '')
            dir_days = u.get('pm_day_count', 0)
            dir_date_range = u.get('pm_date_range', '')
        elif current_direction == 'exc':
            dir_min = u.get('exc_per_min')
            dir_hours_str = u.get('exc_hours_str', '')
            dir_days = u.get('exc_day_count', 0)
            dir_date_range = u.get('exc_date_range', '')
        else:
            dir_min = u.get('am_min') or u.get('pm_min')
            dir_hours_str = u.get('am_hours_str') or u.get('pm_hours_str') or ''
            dir_days = u.get('day_count', 0)
            dir_date_range = u.get('date_range', '')

        # --- 교체 로직 (월 교체는 위에서 이미 처리) ---

        # 2a. 날짜 범위 교체
        if old_date_pattern.match(stripped):
            if dir_days > 0:
                t_el.text = dir_date_range
            else:
                t_el.text = ''

        # 2b. 단독 날짜 교체
        elif re.match(rf'^{old_month_str}\.\d{{2}}$', stripped):
            if dir_days > 0:
                t_el.text = dir_date_range
            else:
                t_el.text = ''

        # 3a. 산출내역 분×일 ("30분×16일", "30 ×17일" 공백 포함)
        elif re.match(r'^\d+분?\s*×\d+일$', stripped):
            if dir_min and dir_days > 0:
                t_el.text = f"{dir_min}분×{dir_days}일"
            else:
                t_el.text = ''

        # 3b. 산출내역 비정형 ("30×1일" - 분 없는 패턴)
        elif re.match(r'^\d+×\d+일$', stripped):
            if dir_min and dir_days > 0:
                t_el.text = f"{dir_min}분×{dir_days}일"
            else:
                t_el.text = ''

        # 4a. 송영시간 "8시간30분" (시간+분 복합)
        elif re.match(r'^\d+시간\d+분$', stripped):
            prev_texts = [
                (all_t[k].text or '').strip()
                for k in range(max(0, i - 5), i)
                if (all_t[k].text or '').strip()
            ]
            if '총 송영서비스 이용시간' in prev_texts:
                t_el.text = f"{u['total_time_str']}"
            elif '제공시간' in prev_texts or '바우처' in prev_texts:
                t_el.text = f"{u['total_time_str']}"
            elif dir_hours_str:
                t_el.text = dir_hours_str
            else:
                t_el.text = ''

        # 4b. 송영시간 분 단독 ("30분")
        elif re.match(r'^\d+분$', stripped):
            prev_texts = [
                (all_t[k].text or '').strip()
                for k in range(max(0, i - 5), i)
                if (all_t[k].text or '').strip()
            ]
            if any('×' in p or '일' in p for p in prev_texts):
                if dir_hours_str:
                    t_el.text = dir_hours_str
                else:
                    t_el.text = ''

        # 4c. 송영시간 "8시간" (정수 시간)
        elif re.match(r'^\d+시간$', stripped):
            prev_texts = [
                (all_t[k].text or '').strip()
                for k in range(max(0, i - 5), i)
                if (all_t[k].text or '').strip()
            ]
            if '총 송영서비스 이용시간' in prev_texts:
                t_el.text = f"{u['total_time_str']}"
            elif '제공시간' in prev_texts or '바우처' in prev_texts:
                t_el.text = f"{u['total_time_str']}"
            elif dir_hours_str:
                t_el.text = dir_hours_str
            else:
                t_el.text = ''

        # 5a. 산출내역 비용 통합 ("9시간×13,820원", "16시간30분×13,820원")
        elif re.match(r'^\d+시간(?:\d+분)?×[\d,]+원$', stripped):
            t_el.text = f"{u['total_time_str']}×{u['unit_price']:,}원"

        # 5b. 산출내역 비용 (원 없음)
        elif re.match(r'^\d+시간(?:\d+분)?×[\d,]*$', stripped):
            t_el.text = f"{u['total_time_str']}×{u['unit_price']:,}원"

        # 5c. 산출내역 비용 분리형 ("×25,910원")
        elif re.match(r'^×[\d,]+원$', stripped):
            t_el.text = f"×{u['unit_price']:,}원"

        # 5d. 산출내역 비용 분리형 (원 없음)
        elif re.match(r'^×[\d,]+$', stripped) and len(stripped) > 3:
            t_el.text = f"×{u['unit_price']:,}원"

        # 6a. 비용 금액 ("414,560원")
        elif re.match(r'^[\d,]+원$', stripped) and len(stripped) > 4:
            t_el.text = f"{u['final_cost']:,}원"

        # 6b. 비용 금액 원 없음 ("124,380" - 콤마 포함 숫자만)
        elif re.match(r'^[\d,]+$', stripped) and ',' in stripped and len(stripped) >= 4:
            t_el.text = f"{u['final_cost']:,}원"

        # 7. 합계 ("=414,560원")
        elif re.match(r'^=[\d,]+원$', stripped):
            t_el.text = f"={u['final_cost']:,}원"

        i += 1

    # 엑셀에는 있지만 HWPX에 없는 이용자 확인
    found_users = set()
    for t_el in all_t:
        txt = (t_el.text or '').strip()
        if txt in users_data:
            found_users.add(txt)

    missing_users = set(users_data.keys()) - found_users
    warnings = []
    if missing_users:
        for name in sorted(missing_users):
            msg = f"⚠️  경고: '{name}' - 엑셀에는 있지만 HWPX 문서에 없습니다!"
            print(msg)
            warnings.append(msg)

    tree.write(section_path, pretty_print=True, xml_declaration=True, encoding='UTF-8')
    return warnings


def generate(hwpx_path, excel_path, output_path, trip_info_path=None):
    """메인 생성 함수"""
    trip_info = {}
    if trip_info_path:
        print("송영정보 파싱 중...")
        trip_info = parse_trip_info(trip_info_path)
        print(f"  {len(trip_info)}명")

    print("엑셀 파일 파싱 중...")
    users_data, new_month = parse_excel(excel_path, trip_info)
    print(f"  {len(users_data)}명, {new_month}월")

    with tempfile.TemporaryDirectory() as tmpdir:
        work = os.path.join(tmpdir, "work")

        print("원본 HWPX unpack...")
        with ZipFile(hwpx_path, 'r') as zf:
            zf.extractall(work)

        section_path = os.path.join(work, "Contents", "section0.xml")
        print("텍스트 교체 중...")
        warnings = replace_texts_in_section(section_path, users_data, new_month) or []

        print("HWPX pack...")
        mimetype_file = os.path.join(work, "mimetype")
        all_files = []
        for root_dir, dirs, files in os.walk(work):
            for fname in sorted(files):
                fpath = os.path.join(root_dir, fname)
                rel = os.path.relpath(fpath, work)
                all_files.append(rel)

        with ZipFile(output_path, 'w', ZIP_DEFLATED) as zf:
            zf.write(mimetype_file, "mimetype", compress_type=ZIP_STORED)
            for rel in all_files:
                if rel == "mimetype":
                    continue
                zf.write(os.path.join(work, rel), rel, compress_type=ZIP_DEFLATED)

        print(f"완료: {output_path}")
    return warnings


def main():
    import argparse
    parser = argparse.ArgumentParser(description='HWPX 텍스트 교체 송영서비스 생성기')
    parser.add_argument('hwpx', help='원본 HWPX 파일 경로')
    parser.add_argument('excel', help='3월 송영서비스 엑셀 파일 경로')
    parser.add_argument('-t', '--trip-info', default=None, help='송영정보 엑셀 파일 경로')
    parser.add_argument('-o', '--output', default=None, help='출력 파일 경로')
    args = parser.parse_args()
    if not args.output:
        args.output = f'주간활동송영서비스_26.{datetime.now().month:02d}.hwpx'
    generate(args.hwpx, args.excel, args.output, args.trip_info)


if __name__ == '__main__':
    main()
